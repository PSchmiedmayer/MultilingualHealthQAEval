#
# This source file is part of the multilingual medical QA evaluation project
#
# SPDX-FileCopyrightText: 2026 Stanford University and the project authors (see CONTRIBUTORS.md)
#
# SPDX-License-Identifier: MIT
#

from __future__ import annotations

import asyncio
import math
import os
import time
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

import pandas as pd
import requests
from requests.exceptions import HTTPError, ReadTimeout


@dataclass(frozen=True)
class ModelConfig:
    name: str
    model_id: str
    temperature: float = 0.0
    max_tokens: int | None = None
    timeout_seconds: int = 180
    max_retries: int = 3
    concurrency: int = 4
    extra_body: dict[str, Any] = field(default_factory=dict)


DEFAULT_TRANSLATION_SYSTEM_PROMPT = (
    "Translate accurately and return only the translated question."
)
DEFAULT_FALLBACK_ANSWER_OUTPUT_TOKENS = 600.0
DEFAULT_OPENROUTER_LEGACY_PRICE_OVERRIDES: dict[str, dict[str, Any]] = {
    "anthropic/claude-3.5-sonnet": {
        "model_name": "Anthropic: Claude 3.5 Sonnet",
        "input_cost_per_million": 6.0,
        "output_cost_per_million": 30.0,
        "pricing_source": "openrouter_legacy_override",
    },
    "google/gemini-pro-1.5": {
        "model_name": "Google: Gemini 1.5 Pro",
        "input_cost_per_million": 1.25,
        "output_cost_per_million": 5.0,
        "pricing_source": "openrouter_legacy_override",
    },
    "meta-llama/llama-3.1-405b-instruct": {
        "model_name": "Meta: Llama 3.1 405B Instruct",
        "input_cost_per_million": 4.0,
        "output_cost_per_million": 4.0,
        "pricing_source": "openrouter_legacy_override",
    },
}


def questions_to_dataframe(questions: Iterable[str | dict[str, Any]]) -> pd.DataFrame:
    question_list = list(questions)
    if not question_list:
        return pd.DataFrame(
            columns=[
                "question_id",
                "question_text",
                "source_language_code",
                "source_language_name",
                "reference_answer",
                "risk_level",
                "topic",
            ]
        )

    if all(isinstance(question, str) for question in question_list):
        frame = pd.DataFrame(
            {
                "question_id": [f"q{i}" for i in range(1, len(question_list) + 1)],
                "question_text": question_list,
            }
        )
    else:
        frame = pd.DataFrame(question_list)
        required = {"question_text"}
        missing = required - set(frame.columns)
        if missing:
            missing_keys = ", ".join(sorted(missing))
            raise ValueError(f"Question entries are missing required keys: {missing_keys}")
        if "question_id" not in frame.columns:
            frame["question_id"] = [f"q{i}" for i in range(1, len(frame) + 1)]

    frame = frame.copy()
    frame["source_language_code"] = frame.get("source_language_code", "en")
    frame["source_language_name"] = frame.get("source_language_name", "English")
    frame["reference_answer"] = frame.get("reference_answer", "")
    frame["risk_level"] = frame.get("risk_level", "")
    frame["topic"] = frame.get("topic", "")
    return frame


def languages_to_dataframe(languages: Iterable[dict[str, Any]]) -> pd.DataFrame:
    frame = pd.DataFrame(languages)
    required = {"language_code", "language_name"}
    missing = required - set(frame.columns)
    if missing:
        missing_keys = ", ".join(sorted(missing))
        raise ValueError(f"Language entries are missing required keys: {missing_keys}")
    return frame.copy()


def review_schema_to_dataframe(schema: Iterable[dict[str, str]]) -> pd.DataFrame:
    frame = pd.DataFrame(schema)
    required = {"field_key", "field_label", "options"}
    missing = required - set(frame.columns)
    if missing:
        raise ValueError(f"Review schema is missing required keys: {', '.join(sorted(missing))}")
    return frame.copy()


def save_authoring_inputs(
    *,
    questions: pd.DataFrame,
    languages: pd.DataFrame,
    clinician_schema: pd.DataFrame,
    lay_schema: pd.DataFrame,
    output_dir: str | Path,
) -> None:
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    questions.to_csv(output_path / "questions.csv", index=False)
    languages.to_csv(output_path / "languages.csv", index=False)
    clinician_schema.to_csv(output_path / "clinician_review_schema.csv", index=False)
    lay_schema.to_csv(output_path / "lay_review_schema.csv", index=False)


def build_question_language_matrix(
    questions: pd.DataFrame,
    languages: pd.DataFrame,
) -> pd.DataFrame:
    left = questions.copy()
    right = languages.copy()
    left["join_key"] = 1
    right["join_key"] = 1
    return left.merge(right, on="join_key").drop(columns=["join_key"])


def _auto_fit_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        lengths = [len(str(cell.value or "")) for cell in column_cells[:25]]
        width = min(max(lengths + [12]) + 2, 40)
        worksheet.column_dimensions[column_cells[0].column_letter].width = width


def _style_and_protect_sheet(worksheet, editable_columns: set[str]) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill, Protection

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    editable_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    header_map = {cell.value: cell.column for cell in worksheet[1]}
    editable_indexes = {header_map[name] for name in editable_columns if name in header_map}

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.protection = Protection(locked=True)
            if cell.column in editable_indexes:
                cell.protection = Protection(locked=False)
                cell.fill = editable_fill

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    _auto_fit_columns(worksheet)
    worksheet.protection.sheet = True


def _add_dropdown_validation(
    workbook,
    worksheet_name: str,
    column_name: str,
    options: list[str],
) -> None:
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    if worksheet_name not in workbook.sheetnames or "_lists" not in workbook.sheetnames:
        return

    worksheet = workbook[worksheet_name]
    header_map = {cell.value: cell.column_letter for cell in worksheet[1]}
    target_column = header_map.get(column_name)
    if not target_column:
        return

    lists_sheet = workbook["_lists"]
    list_column = lists_sheet.max_column + 1
    for index, option in enumerate(options, start=1):
        lists_sheet.cell(row=index, column=list_column, value=option)

    column_letter = get_column_letter(list_column)
    formula = f"'_lists'!${column_letter}$1:${column_letter}${len(options)}"

    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.prompt = f"Select a value for {column_name}"
    validation.error = "Please choose one of the dropdown options."
    worksheet.add_data_validation(validation)
    validation.add(f"{target_column}2:{target_column}1048576")


def _finalize_workbook(
    path: Path,
    worksheet_name: str,
    editable_columns: set[str],
    dropdown_map: dict[str, list[str]],
) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    if "_lists" not in workbook.sheetnames:
        workbook.create_sheet("_lists")
    workbook["_lists"].sheet_state = "hidden"

    _style_and_protect_sheet(workbook[worksheet_name], editable_columns)
    if "instructions" in workbook.sheetnames:
        _style_and_protect_sheet(workbook["instructions"], set())

    for column_name, options in dropdown_map.items():
        _add_dropdown_validation(workbook, worksheet_name, column_name, options)

    workbook.save(path)


def _load_checkpoint(path: str | Path | None) -> pd.DataFrame:
    if path is None:
        return pd.DataFrame()
    checkpoint_path = Path(path)
    if not checkpoint_path.exists():
        return pd.DataFrame()
    return load_table(checkpoint_path)


def load_table(path: str | Path) -> pd.DataFrame:
    table_path = Path(path)
    if table_path.suffix == ".parquet":
        return pd.read_parquet(table_path)
    return pd.read_csv(table_path)


def save_table(frame: pd.DataFrame, path: str | Path) -> None:
    table_path = Path(path)
    table_path.parent.mkdir(parents=True, exist_ok=True)
    if table_path.suffix == ".parquet":
        frame.to_parquet(table_path, index=False)
    else:
        frame.to_csv(table_path, index=False)


def _save_checkpoint(frame: pd.DataFrame, path: str | Path | None) -> None:
    if path is None:
        return
    save_table(frame, path)


def _format_eta(seconds: float | None) -> str:
    if seconds is None:
        return "unknown"
    if seconds < 60:
        return f"{seconds:.0f}s"
    minutes, remaining_seconds = divmod(int(seconds), 60)
    return f"{minutes}m {remaining_seconds}s"


def _estimate_remaining_seconds(
    *,
    elapsed_seconds: float,
    completed_in_run: int,
    remaining_in_run: int,
    concurrency: int,
) -> float | None:
    if completed_in_run <= 0 or elapsed_seconds <= 0 or remaining_in_run <= 0:
        return None

    parallelism = max(1, concurrency)
    throughput_seconds = remaining_in_run / (completed_in_run / elapsed_seconds)
    batch_seconds = elapsed_seconds / math.ceil(completed_in_run / parallelism)
    batch_eta = math.ceil(remaining_in_run / parallelism) * batch_seconds

    return min(throughput_seconds, batch_eta)


async def _call_openrouter_async(
    *,
    model: ModelConfig,
    prompt: str,
    system_prompt: str | None = None,
) -> str:
    return await asyncio.to_thread(
        _call_openrouter,
        model=model,
        prompt=prompt,
        system_prompt=system_prompt,
    )


def _openrouter_headers() -> dict[str, str]:
    api_key = os.getenv("OPENROUTER_API_KEY")
    if not api_key:
        raise EnvironmentError("OPENROUTER_API_KEY is not set.")
    return {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }


def _call_openrouter(
    *,
    model: ModelConfig,
    prompt: str,
    system_prompt: str | None = None,
) -> str:
    messages: list[dict[str, str]] = []
    if system_prompt:
        messages.append({"role": "system", "content": system_prompt})
    messages.append({"role": "user", "content": prompt})

    payload: dict[str, Any] = {
        "model": model.model_id,
        "messages": messages,
        "temperature": model.temperature,
        "extra_body": {
            "data_collection": "deny",
            "zdr": True,
            **model.extra_body,
        },
    }
    if model.max_tokens is not None:
        payload["max_tokens"] = model.max_tokens

    last_error: Exception | None = None
    for attempt in range(1, model.max_retries + 1):
        try:
            response = requests.post(
                "https://openrouter.ai/api/v1/chat/completions",
                headers=_openrouter_headers(),
                json=payload,
                timeout=model.timeout_seconds,
            )
            response.raise_for_status()
            body = response.json()
            return body["choices"][0]["message"]["content"].strip()
        except ReadTimeout as error:
            last_error = error
            if attempt == model.max_retries:
                break
            wait_seconds = min(2 ** (attempt - 1), 8)
            print(
                f"Request timed out for {model.name} "
                f"(attempt {attempt}/{model.max_retries}). Retrying in {wait_seconds}s..."
            )
            time.sleep(wait_seconds)
        except HTTPError as error:
            status_code = error.response.status_code if error.response is not None else None
            if status_code not in {429, 500, 502, 503, 504}:
                raise

            last_error = error
            if attempt == model.max_retries:
                break

            retry_after_header = None
            if error.response is not None:
                retry_after_header = error.response.headers.get("Retry-After")

            if retry_after_header and retry_after_header.isdigit():
                wait_seconds = int(retry_after_header)
            else:
                wait_seconds = min(5 * (2 ** (attempt - 1)), 60)

            print(
                f"Request failed for {model.name} with HTTP {status_code} "
                f"(attempt {attempt}/{model.max_retries}). Retrying in {wait_seconds}s..."
            )
            time.sleep(wait_seconds)

    if isinstance(last_error, HTTPError) and last_error.response is not None:
        status_code = last_error.response.status_code
        raise RuntimeError(
            f"OpenRouter request failed for {model.name} with HTTP {status_code} "
            f"after {model.max_retries} attempts."
        ) from last_error

    raise RuntimeError(
        f"OpenRouter request timed out for {model.name} after {model.max_retries} attempts "
        f"with a {model.timeout_seconds}s timeout."
    ) from last_error


async def propose_translations(
    matrix: pd.DataFrame,
    *,
    translation_model: ModelConfig,
    translation_system_prompt: str | None = (
        "Translate accurately and return only the translated question."
    ),
    show_progress: bool = True,
    checkpoint_path: str | Path | None = None,
    force_restart: bool = False,
) -> pd.DataFrame:
    existing = pd.DataFrame() if force_restart else _load_checkpoint(checkpoint_path)
    completed_rows: list[dict[str, Any]] = []
    completed_lookup: dict[tuple[str, str], dict[str, Any]] = {}
    total = len(matrix)
    if not existing.empty:
        for row in existing.to_dict(orient="records"):
            key = (str(row.get("question_id", "")), str(row.get("language_code", "")))
            completed_lookup[key] = row

    for row in matrix.to_dict(orient="records"):
        key = (str(row.get("question_id", "")), str(row.get("language_code", "")))
        existing_row = completed_lookup.get(key)
        if existing_row and existing_row.get("question_text") == row.get("question_text"):
            completed_rows.append(existing_row)

    resumed_count = len(completed_rows)
    remaining_total = total - resumed_count

    if show_progress and resumed_count:
        print(
            f"Resuming translations from checkpoint: {resumed_count} already saved, "
            f"{remaining_total} remaining."
        )
    if show_progress and remaining_total == 0:
        print(f"Translations already complete: {total}/{total}.")
        return pd.DataFrame(completed_rows)

    started = time.monotonic()
    new_completed = 0
    pending_rows = []
    for row in matrix.itertuples(index=False):
        key = (str(row.question_id), str(row.language_code))
        existing_row = completed_lookup.get(key)
        if existing_row and existing_row.get("question_text") == row.question_text:
            continue
        pending_rows.append(row)

    semaphore = asyncio.Semaphore(max(1, translation_model.concurrency))

    async def translate_one(row):
        async with semaphore:
            source_language_name = getattr(row, "source_language_name", None) or "English"
            if row.source_language_code == row.language_code:
                suggestion = row.question_text
            else:
                prompt = (
                    f"Translate the following medical question from {source_language_name} "
                    f"to {row.language_name}. Preserve meaning, ambiguity, and clinical intent.\n\n"
                    f"Question: {row.question_text}"
                )
                suggestion = await _call_openrouter_async(
                    model=translation_model,
                    prompt=prompt,
                    system_prompt=translation_system_prompt,
                )

            row_dict = row._asdict()
            row_dict["translation_suggestion"] = suggestion
            return row_dict

    tasks = [asyncio.create_task(translate_one(row)) for row in pending_rows]
    for task in asyncio.as_completed(tasks):
        row_dict = await task
        completed_rows.append(row_dict)
        _save_checkpoint(pd.DataFrame(completed_rows), checkpoint_path)
        new_completed += 1

        if show_progress:
            completed = resumed_count + new_completed
            elapsed = time.monotonic() - started
            remaining_seconds = _estimate_remaining_seconds(
                elapsed_seconds=elapsed,
                completed_in_run=new_completed,
                remaining_in_run=remaining_total - new_completed,
                concurrency=translation_model.concurrency,
            )
            print(
                f"Translations: {completed}/{total} complete "
                f"(elapsed {elapsed:.0f}s, eta {_format_eta(remaining_seconds)})"
            )

    ordered_rows = []
    for row in matrix.itertuples(index=False):
        key = (str(row.question_id), str(row.language_code))
        match = next(
            item for item in completed_rows
            if str(item.get("question_id")) == key[0] and str(item.get("language_code")) == key[1]
        )
        ordered_rows.append(match)
    return pd.DataFrame(ordered_rows)


def _translation_export_frame(translations: pd.DataFrame) -> pd.DataFrame:
    export_frame = translations.copy()
    export_frame["translation_key"] = (
        export_frame["question_id"].astype(str) + "::" + export_frame["language_code"].astype(str)
    )
    export_frame["translation_final"] = export_frame["translation_suggestion"]
    export_frame["translation_review_status"] = ""
    export_frame["translation_reviewer"] = ""
    export_frame["translation_notes"] = ""

    preferred_columns = [
        "translation_key",
        "question_id",
        "language_code",
        "language_name",
        "category",
        "topic",
        "risk_level",
        "question_text",
        "reference_answer",
        "notes",
        "translation_suggestion",
        "translation_final",
        "translation_review_status",
        "translation_reviewer",
        "translation_notes",
    ]
    available_columns = [column for column in preferred_columns if column in export_frame.columns]
    return export_frame[available_columns]


def export_translation_review_workbooks(
    translations: pd.DataFrame,
    output_dir: str | Path,
) -> list[Path]:
    output_path = Path(output_dir)
    generated_dir = output_path / "generated"
    completed_dir = output_path / "completed"
    generated_dir.mkdir(parents=True, exist_ok=True)
    completed_dir.mkdir(parents=True, exist_ok=True)

    instructions = pd.DataFrame(
        {
            "step": [
                "Review the suggested translation for one language only.",
                (
                    "Edit only the yellow columns: translation_final, "
                    "translation_review_status, translation_reviewer, "
                    "and translation_notes."
                ),
                "Use translation_review_status values like approved or needs_changes.",
                "Save the completed file into the completed folder for this language group.",
            ]
        }
    )

    export_paths: list[Path] = []
    editable_columns = {
        "translation_final",
        "translation_review_status",
        "translation_reviewer",
        "translation_notes",
    }
    dropdown_map = {
        "translation_review_status": ["approved", "needs_changes"],
    }

    for language_code, language_frame in translations.groupby("language_code", dropna=False):
        export_frame = _translation_export_frame(language_frame)
        file_path = generated_dir / f"translation_review_{language_code}.xlsx"
        with pd.ExcelWriter(file_path) as writer:
            export_frame.to_excel(writer, sheet_name="translations", index=False)
            instructions.to_excel(writer, sheet_name="instructions", index=False)
        _finalize_workbook(file_path, "translations", editable_columns, dropdown_map)
        export_paths.append(file_path)

    return export_paths


def load_reviewed_translations(completed_dir: str | Path) -> pd.DataFrame:
    completed_path = Path(completed_dir)
    if not completed_path.exists():
        return pd.DataFrame()

    frames: list[pd.DataFrame] = []
    for workbook_path in sorted(completed_path.glob("*.xlsx")):
        frame = pd.read_excel(workbook_path, sheet_name="translations")
        frame["translation_review_file"] = workbook_path.name
        frames.append(frame)

    if not frames:
        return pd.DataFrame()

    frame = pd.concat(frames, ignore_index=True)
    final_text = frame["translation_final"].fillna("").astype(str).str.strip()
    frame["translated_question"] = final_text.where(
        final_text != "",
        frame["translation_suggestion"],
    )
    return frame


async def generate_answers(
    translations: pd.DataFrame,
    *,
    models: Iterable[ModelConfig],
    system_prompt: str | None = None,
    repeats_per_prompt: int = 1,
    show_progress: bool = True,
    checkpoint_path: str | Path | None = None,
    force_restart: bool = False,
) -> pd.DataFrame:
    existing = pd.DataFrame() if force_restart else _load_checkpoint(checkpoint_path)
    answers: list[dict[str, Any]] = []
    model_list = list(models)
    total = len(translations) * len(model_list) * repeats_per_prompt
    completed_lookup: dict[tuple[str, str, str, int], dict[str, Any]] = {}
    if not existing.empty:
        for row in existing.to_dict(orient="records"):
            key = (
                str(row.get("question_id", "")),
                str(row.get("language_code", "")),
                str(row.get("model_id", "")),
                int(row.get("repeat_index", 0)),
            )
            completed_lookup[key] = row

    for row in translations.to_dict(orient="records"):
        for model in model_list:
            for repeat_index in range(1, repeats_per_prompt + 1):
                key = (
                    str(row.get("question_id", "")),
                    str(row.get("language_code", "")),
                    model.model_id,
                    repeat_index,
                )
                existing_row = completed_lookup.get(key)
                if (
                    existing_row
                    and existing_row.get("translated_question")
                    == row.get("translated_question")
                ):
                    answers.append(existing_row)

    resumed_count = len(answers)
    remaining_total = total - resumed_count

    if show_progress and resumed_count:
        print(
            f"Resuming answers from checkpoint: {resumed_count} already saved, "
            f"{remaining_total} remaining."
        )
    if show_progress and remaining_total == 0:
        print(f"Answers already complete: {total}/{total}.")
        return pd.DataFrame(answers)

    started = time.monotonic()
    new_completed = 0
    pending_items = []
    for row in translations.itertuples(index=False):
        for model in model_list:
            for repeat_index in range(1, repeats_per_prompt + 1):
                key = (str(row.question_id), str(row.language_code), model.model_id, repeat_index)
                existing_row = completed_lookup.get(key)
                if (
                    existing_row
                    and existing_row.get("translated_question") == row.translated_question
                ):
                    continue
                pending_items.append((row, model, repeat_index))

    max_concurrency = max(1, min(model.concurrency for model in model_list)) if model_list else 1
    semaphore = asyncio.Semaphore(max_concurrency)

    async def answer_one(row, model, repeat_index):
        async with semaphore:
            answer_text = await _call_openrouter_async(
                model=model,
                prompt=row.translated_question,
                system_prompt=system_prompt,
            )
            return {
                **row._asdict(),
                "response_key": (
                    f"{row.question_id}::{row.language_code}::"
                    f"{model.model_id}::{repeat_index}"
                ),
                "candidate_name": model.name,
                "model_id": model.model_id,
                "repeat_index": repeat_index,
                "answer_text": answer_text,
            }

    tasks = [
        asyncio.create_task(answer_one(row, model, repeat_index))
        for row, model, repeat_index in pending_items
    ]
    for task in asyncio.as_completed(tasks):
        answer_row = await task
        answers.append(answer_row)
        _save_checkpoint(pd.DataFrame(answers), checkpoint_path)
        new_completed += 1

        if show_progress:
            completed = resumed_count + new_completed
            elapsed = time.monotonic() - started
            remaining_seconds = _estimate_remaining_seconds(
                elapsed_seconds=elapsed,
                completed_in_run=new_completed,
                remaining_in_run=remaining_total - new_completed,
                concurrency=max_concurrency,
            )
            print(
                f"Answers: {completed}/{total} complete "
                f"(elapsed {elapsed:.0f}s, eta {_format_eta(remaining_seconds)})"
            )

    ordered_rows = []
    for row in translations.itertuples(index=False):
        for model in model_list:
            for repeat_index in range(1, repeats_per_prompt + 1):
                key = (str(row.question_id), str(row.language_code), model.model_id, repeat_index)
                match = next(
                    item for item in answers
                    if (
                        str(item.get("question_id")) == key[0]
                        and str(item.get("language_code")) == key[1]
                        and str(item.get("model_id")) == key[2]
                        and int(item.get("repeat_index")) == key[3]
                    )
                )
                ordered_rows.append(match)
    return pd.DataFrame(ordered_rows)


def _response_export_frame(answers: pd.DataFrame, review_schema: pd.DataFrame) -> pd.DataFrame:
    frame = answers.copy()
    for row in review_schema.itertuples(index=False):
        frame[f"review__{row.field_key}"] = ""
    frame["reviewer_name"] = ""
    frame["review_status"] = ""
    frame["review_notes"] = ""

    preferred_columns = [
        "response_key",
        "question_id",
        "language_code",
        "language_name",
        "category",
        "topic",
        "risk_level",
        "candidate_name",
        "model_id",
        "repeat_index",
        "question_text",
        "translated_question",
        "reference_answer",
        "answer_text",
        "reviewer_name",
        "review_status",
    ]
    review_columns = [f"review__{row.field_key}" for row in review_schema.itertuples(index=False)]
    trailing_columns = ["review_notes"]
    ordered = [column for column in preferred_columns if column in frame.columns]
    ordered += [column for column in review_columns if column in frame.columns]
    ordered += [column for column in trailing_columns if column in frame.columns]
    remaining = [column for column in frame.columns if column not in ordered]
    return frame[ordered + remaining]


def export_review_workbooks(
    answers: pd.DataFrame,
    *,
    review_schema: pd.DataFrame,
    output_dir: str | Path,
    review_type: str,
) -> list[Path]:
    output_path = Path(output_dir)
    generated_dir = output_path / "generated"
    completed_dir = output_path / "completed"
    generated_dir.mkdir(parents=True, exist_ok=True)
    completed_dir.mkdir(parents=True, exist_ok=True)

    response_frame = _response_export_frame(answers, review_schema)
    instructions = pd.DataFrame(
        {
            "step": [
                f"Review the {review_type} rubric columns in the responses sheet.",
                (
                    "Each row has a stable response_key so multiple reviewers "
                    "can assess the same answer independently."
                ),
                (
                    "Enter your name in reviewer_name and use the dropdown "
                    "fields for the rubric selections."
                ),
                "Leave the original question, translated question, and answer text unchanged.",
                "Save completed files into the completed folder so the notebook can import them.",
            ]
        }
    )

    export_paths: list[Path] = []
    editable_columns = {"reviewer_name", "review_status", "review_notes"}
    dropdown_map = {"review_status": ["completed", "partial", "needs_follow_up"]}
    for row in review_schema.itertuples(index=False):
        column_name = f"review__{row.field_key}"
        editable_columns.add(column_name)
        dropdown_map[column_name] = [option.strip() for option in row.options.split("|")]

    master_path = generated_dir / f"{review_type}_all_languages.xlsx"
    with pd.ExcelWriter(master_path) as writer:
        response_frame.to_excel(writer, sheet_name="responses", index=False)
        review_schema.to_excel(writer, sheet_name="criteria", index=False)
        instructions.to_excel(writer, sheet_name="instructions", index=False)
    _finalize_workbook(master_path, "responses", editable_columns, dropdown_map)
    export_paths.append(master_path)

    for language_code, language_frame in response_frame.groupby("language_code", dropna=False):
        file_path = generated_dir / f"{review_type}_{language_code}.xlsx"
        with pd.ExcelWriter(file_path) as writer:
            language_frame.to_excel(writer, sheet_name="responses", index=False)
            review_schema.to_excel(writer, sheet_name="criteria", index=False)
            instructions.to_excel(writer, sheet_name="instructions", index=False)
        _finalize_workbook(file_path, "responses", editable_columns, dropdown_map)
        export_paths.append(file_path)

    return export_paths


def load_completed_reviews(completed_dir: str | Path) -> pd.DataFrame:
    completed_path = Path(completed_dir)
    if not completed_path.exists():
        return pd.DataFrame()

    review_frames: list[pd.DataFrame] = []
    for workbook in sorted(completed_path.glob("*.xlsx")):
        response_frame = pd.read_excel(workbook, sheet_name="responses")
        response_frame["review_file"] = workbook.name
        review_frames.append(response_frame)

    if not review_frames:
        return pd.DataFrame()

    return pd.concat(review_frames, ignore_index=True)


def summarize_review_answers(
    reviews: pd.DataFrame,
    *,
    review_schema: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    if reviews.empty:
        return pd.DataFrame(), pd.DataFrame()

    reviews = reviews.copy()
    review_columns = [f"review__{field_key}" for field_key in review_schema["field_key"]]
    for column in review_columns:
        if column not in reviews.columns:
            reviews[column] = ""

    id_vars = [
        "candidate_name",
        "model_id",
        "language_code",
        "language_name",
        "question_id",
        "repeat_index",
        "review_file",
    ]
    available_id_vars = [column for column in id_vars if column in reviews.columns]

    long_reviews = reviews.melt(
        id_vars=available_id_vars,
        value_vars=review_columns,
        var_name="review_field",
        value_name="selected_option",
    )
    long_reviews = long_reviews.dropna(subset=["selected_option"])
    long_reviews["selected_option"] = long_reviews["selected_option"].astype(str).str.strip()
    long_reviews = long_reviews[long_reviews["selected_option"] != ""]
    long_reviews["field_key"] = long_reviews["review_field"].str.removeprefix("review__")
    long_reviews = long_reviews.merge(review_schema, on="field_key", how="left")

    summary = (
        long_reviews.groupby(
            ["candidate_name", "language_code", "field_key", "field_label", "selected_option"],
            dropna=False,
        )
        .size()
        .reset_index(name="count")
    )
    totals = (
        summary.groupby(["candidate_name", "language_code", "field_key"], dropna=False)["count"]
        .sum()
        .reset_index(name="total_count")
    )
    summary = summary.merge(
        totals,
        on=["candidate_name", "language_code", "field_key"],
        how="left",
    )
    summary["share"] = summary["count"] / summary["total_count"]
    return long_reviews, summary


def model_config_records(models: Iterable[ModelConfig]) -> pd.DataFrame:
    return pd.DataFrame([asdict(model) for model in models])


def estimate_token_count(text: str | None, *, chars_per_token: float = 4.0) -> int:
    normalized = str(text or "").strip()
    if not normalized:
        return 0
    safe_chars_per_token = max(chars_per_token, 0.1)
    return max(1, math.ceil(len(normalized) / safe_chars_per_token))


def _first_available_token_average(
    frame: pd.DataFrame | None,
    columns: Sequence[str],
    *,
    chars_per_token: float,
) -> tuple[str | None, float | None]:
    if frame is None or frame.empty:
        return None, None

    for column in columns:
        if column not in frame.columns:
            continue
        series = frame[column].fillna("").astype(str).str.strip()
        series = series[series != ""]
        if series.empty:
            continue
        token_average = series.map(
            lambda value: estimate_token_count(value, chars_per_token=chars_per_token)
        ).mean()
        return column, float(token_average)

    return None, None


def fetch_openrouter_model_catalog(*, timeout_seconds: int = 30) -> pd.DataFrame:
    headers = {"Content-Type": "application/json"}
    api_key = os.getenv("OPENROUTER_API_KEY")
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"

    response = requests.get(
        "https://openrouter.ai/api/v1/models",
        headers=headers,
        timeout=timeout_seconds,
    )
    response.raise_for_status()

    records: list[dict[str, Any]] = []
    for item in response.json().get("data", []):
        pricing = item.get("pricing") or {}
        prompt_cost = float(pricing.get("prompt") or 0.0) * 1_000_000
        completion_cost = float(pricing.get("completion") or 0.0) * 1_000_000
        records.append(
            {
                "model_id": str(item.get("id", "")),
                "model_name": str(item.get("name", item.get("id", ""))),
                "created": int(item.get("created", 0) or 0),
                "input_cost_per_million": prompt_cost,
                "output_cost_per_million": completion_cost,
                "pricing_source": "openrouter_catalog",
            }
        )

    frame = pd.DataFrame(records)
    if frame.empty:
        return frame
    return frame.sort_values(
        ["created", "model_id"],
        ascending=[False, True],
    ).reset_index(drop=True)


def resolve_openrouter_model_pricing(
    model_ids: Iterable[str],
    *,
    catalog: pd.DataFrame | None = None,
    price_overrides: Mapping[str, Mapping[str, Any]] | None = None,
    timeout_seconds: int = 30,
) -> pd.DataFrame:
    unique_model_ids = list(dict.fromkeys(str(model_id) for model_id in model_ids if str(model_id)))
    if not unique_model_ids:
        return pd.DataFrame(
            columns=[
                "model_id",
                "model_name",
                "created",
                "input_cost_per_million",
                "output_cost_per_million",
                "pricing_source",
            ]
        )

    catalog_frame = (
        fetch_openrouter_model_catalog(timeout_seconds=timeout_seconds)
        if catalog is None
        else catalog
    )
    catalog_lookup: dict[str, dict[str, Any]] = {}
    if not catalog_frame.empty:
        for row in catalog_frame.to_dict(orient="records"):
            model_id = str(row.get("model_id", ""))
            if model_id in unique_model_ids and model_id not in catalog_lookup:
                catalog_lookup[model_id] = row

    merged_overrides: dict[str, Mapping[str, Any]] = dict(DEFAULT_OPENROUTER_LEGACY_PRICE_OVERRIDES)
    if price_overrides:
        merged_overrides.update(
            {
                str(model_id): override
                for model_id, override in price_overrides.items()
            }
        )

    resolved: dict[str, dict[str, Any]] = {}
    for model_id in unique_model_ids:
        if model_id in catalog_lookup:
            resolved[model_id] = dict(catalog_lookup[model_id])

    for model_id, override in merged_overrides.items():
        if model_id not in unique_model_ids:
            continue
        resolved[model_id] = {
            "model_id": model_id,
            "model_name": str(override.get("model_name", model_id)),
            "created": int(override.get("created", 0) or 0),
            "input_cost_per_million": float(override.get("input_cost_per_million", 0.0) or 0.0),
            "output_cost_per_million": float(override.get("output_cost_per_million", 0.0) or 0.0),
            "pricing_source": str(override.get("pricing_source", "override")),
        }

    missing = [model_id for model_id in unique_model_ids if model_id not in resolved]
    if missing:
        missing_models = ", ".join(sorted(missing))
        raise ValueError(f"Missing OpenRouter pricing for model IDs: {missing_models}")

    ordered_rows = [resolved[model_id] for model_id in unique_model_ids]
    return pd.DataFrame(ordered_rows)


def _build_cost_assumptions(
    *,
    questions: pd.DataFrame,
    translation_samples: pd.DataFrame | None,
    answer_samples: pd.DataFrame | None,
    translation_system_prompt: str | None,
    answer_system_prompt: str | None,
    chars_per_token: float,
    fallback_answer_output_tokens: float,
) -> tuple[dict[str, Any], pd.DataFrame]:
    _, question_tokens = _first_available_token_average(
        questions,
        ["question_text"],
        chars_per_token=chars_per_token,
    )
    question_tokens = float(question_tokens or 0.0)

    translated_source, translated_tokens = _first_available_token_average(
        translation_samples,
        ["translated_question", "translation_suggestion"],
        chars_per_token=chars_per_token,
    )
    if translated_tokens is None:
        translated_tokens = question_tokens
        translated_source = "question_text_fallback"

    answer_output_source, answer_output_tokens = _first_available_token_average(
        answer_samples,
        ["answer_text"],
        chars_per_token=chars_per_token,
    )
    if answer_output_tokens is None:
        answer_output_tokens = float(fallback_answer_output_tokens)
        answer_output_source = "fallback_default"

    translation_system_tokens = float(
        estimate_token_count(translation_system_prompt, chars_per_token=chars_per_token)
    )
    answer_system_tokens = float(
        estimate_token_count(answer_system_prompt, chars_per_token=chars_per_token)
    )

    assumptions = {
        "chars_per_token": float(chars_per_token),
        "question_tokens_per_request": question_tokens,
        "question_tokens_source": "question_text",
        "translated_question_tokens_per_request": float(translated_tokens),
        "translated_question_tokens_source": str(translated_source),
        "translation_prompt_tokens_per_request": question_tokens + translation_system_tokens,
        "translation_prompt_tokens_source": "question_text + translation_system_prompt",
        "translation_completion_tokens_per_request": float(translated_tokens),
        "translation_completion_tokens_source": str(translated_source),
        "answer_prompt_tokens_per_request": float(translated_tokens) + answer_system_tokens,
        "answer_prompt_tokens_source": "translated_question + answer_system_prompt",
        "answer_completion_tokens_per_request": float(answer_output_tokens),
        "answer_completion_tokens_source": str(answer_output_source),
        "translation_system_prompt_tokens": translation_system_tokens,
        "answer_system_prompt_tokens": answer_system_tokens,
    }

    assumption_rows = [
        {
            "metric": "chars_per_token",
            "estimated_tokens": assumptions["chars_per_token"],
            "source": "heuristic",
        },
        {
            "metric": "question_tokens_per_request",
            "estimated_tokens": assumptions["question_tokens_per_request"],
            "source": assumptions["question_tokens_source"],
        },
        {
            "metric": "translated_question_tokens_per_request",
            "estimated_tokens": assumptions["translated_question_tokens_per_request"],
            "source": assumptions["translated_question_tokens_source"],
        },
        {
            "metric": "translation_prompt_tokens_per_request",
            "estimated_tokens": assumptions["translation_prompt_tokens_per_request"],
            "source": assumptions["translation_prompt_tokens_source"],
        },
        {
            "metric": "translation_completion_tokens_per_request",
            "estimated_tokens": assumptions["translation_completion_tokens_per_request"],
            "source": assumptions["translation_completion_tokens_source"],
        },
        {
            "metric": "answer_prompt_tokens_per_request",
            "estimated_tokens": assumptions["answer_prompt_tokens_per_request"],
            "source": assumptions["answer_prompt_tokens_source"],
        },
        {
            "metric": "answer_completion_tokens_per_request",
            "estimated_tokens": assumptions["answer_completion_tokens_per_request"],
            "source": assumptions["answer_completion_tokens_source"],
        },
    ]
    assumptions_frame = pd.DataFrame(assumption_rows)
    assumptions_frame["estimated_tokens"] = assumptions_frame["estimated_tokens"].round(1)
    return assumptions, assumptions_frame


def estimate_openrouter_cost_scenarios(
    *,
    questions: pd.DataFrame,
    languages: pd.DataFrame,
    scenarios: Mapping[str, Sequence[str]],
    repeats_per_prompt: int = 1,
    translation_model_id: str | None = None,
    question_count: int | None = None,
    language_count: int | None = None,
    translation_samples: pd.DataFrame | None = None,
    answer_samples: pd.DataFrame | None = None,
    translation_system_prompt: str | None = DEFAULT_TRANSLATION_SYSTEM_PROMPT,
    answer_system_prompt: str | None = None,
    price_overrides: Mapping[str, Mapping[str, Any]] | None = None,
    chars_per_token: float = 4.0,
    fallback_answer_output_tokens: float = DEFAULT_FALLBACK_ANSWER_OUTPUT_TOKENS,
    timeout_seconds: int = 30,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    normalized_scenarios = {
        str(name): [str(model_id) for model_id in model_ids if str(model_id)]
        for name, model_ids in scenarios.items()
    }

    all_model_ids: list[str] = []
    for model_ids in normalized_scenarios.values():
        all_model_ids.extend(model_ids)
    if translation_model_id:
        all_model_ids.append(str(translation_model_id))

    pricing_frame = resolve_openrouter_model_pricing(
        all_model_ids,
        price_overrides=price_overrides,
        timeout_seconds=timeout_seconds,
    )
    pricing_lookup = {
        str(row["model_id"]): row for row in pricing_frame.to_dict(orient="records")
    }

    assumptions, assumptions_frame = _build_cost_assumptions(
        questions=questions,
        translation_samples=translation_samples,
        answer_samples=answer_samples,
        translation_system_prompt=translation_system_prompt,
        answer_system_prompt=answer_system_prompt,
        chars_per_token=chars_per_token,
        fallback_answer_output_tokens=fallback_answer_output_tokens,
    )

    scenario_question_count = int(
        question_count if question_count is not None else len(questions)
    )
    scenario_language_count = int(
        language_count if language_count is not None else len(languages)
    )
    requests_per_model = (
        scenario_question_count
        * scenario_language_count
        * max(1, repeats_per_prompt)
    )
    translation_request_count = (
        scenario_question_count * scenario_language_count if translation_model_id else 0
    )

    translation_cost = 0.0
    translation_price_source = ""
    if translation_model_id:
        translation_price = pricing_lookup[str(translation_model_id)]
        translation_input_cost = (
            translation_request_count
            * assumptions["translation_prompt_tokens_per_request"]
            * float(translation_price["input_cost_per_million"])
            / 1_000_000
        )
        translation_output_cost = (
            translation_request_count
            * assumptions["translation_completion_tokens_per_request"]
            * float(translation_price["output_cost_per_million"])
            / 1_000_000
        )
        translation_cost = translation_input_cost + translation_output_cost
        translation_price_source = str(translation_price["pricing_source"])

    summary_rows: list[dict[str, Any]] = []
    per_model_rows: list[dict[str, Any]] = []
    for scenario_name, model_ids in normalized_scenarios.items():
        answer_cost_total = 0.0
        for model_id in model_ids:
            price = pricing_lookup[model_id]
            estimated_input_cost = (
                requests_per_model
                * assumptions["answer_prompt_tokens_per_request"]
                * float(price["input_cost_per_million"])
                / 1_000_000
            )
            estimated_output_cost = (
                requests_per_model
                * assumptions["answer_completion_tokens_per_request"]
                * float(price["output_cost_per_million"])
                / 1_000_000
            )
            estimated_total_cost = estimated_input_cost + estimated_output_cost
            answer_cost_total += estimated_total_cost
            per_model_rows.append(
                {
                    "scenario_name": scenario_name,
                    "model_id": model_id,
                    "model_name": price["model_name"],
                    "pricing_source": price["pricing_source"],
                    "request_count": requests_per_model,
                    "input_cost_per_million": float(price["input_cost_per_million"]),
                    "output_cost_per_million": float(price["output_cost_per_million"]),
                    "estimated_input_cost_usd": estimated_input_cost,
                    "estimated_output_cost_usd": estimated_output_cost,
                    "estimated_total_cost_usd": estimated_total_cost,
                }
            )

        summary_rows.append(
            {
                "scenario_name": scenario_name,
                "question_count": scenario_question_count,
                "language_count": scenario_language_count,
                "model_count": len(model_ids),
                "repeats_per_prompt": max(1, repeats_per_prompt),
                "translation_model_id": str(translation_model_id or ""),
                "translation_pricing_source": translation_price_source,
                "translation_request_count": translation_request_count,
                "answer_request_count": requests_per_model * len(model_ids),
                "translation_estimated_cost_usd": translation_cost,
                "answer_estimated_cost_usd": answer_cost_total,
                "total_estimated_cost_usd": translation_cost + answer_cost_total,
            }
        )

    summary_frame = pd.DataFrame(summary_rows)
    per_model_frame = pd.DataFrame(per_model_rows)

    for frame in [summary_frame, per_model_frame]:
        for column in frame.columns:
            if column.endswith("_usd") or column.endswith("_per_million"):
                frame[column] = frame[column].round(4)

    return assumptions_frame, summary_frame, per_model_frame


def build_budget_recommendations(
    cost_summary: pd.DataFrame,
    *,
    extra_answer_iterations: Sequence[int] = (0, 1, 2, 3),
    round_up_to_usd: float = 5.0,
    recommended_extra_iterations: int = 2,
) -> pd.DataFrame:
    if cost_summary.empty:
        return pd.DataFrame(
            columns=[
                "scenario_name",
                "budget_tier",
                "extra_answer_iterations",
                "target_repeats_per_prompt",
                "estimated_total_cost_usd",
                "communicated_budget_usd",
                "recommended_to_communicate",
            ]
        )

    safe_round_increment = max(float(round_up_to_usd), 0.01)
    rows: list[dict[str, Any]] = []
    for row in cost_summary.to_dict(orient="records"):
        base_repeats = max(1, int(row.get("repeats_per_prompt", 1) or 1))
        translation_cost = float(row.get("translation_estimated_cost_usd", 0.0) or 0.0)
        answer_cost = float(row.get("answer_estimated_cost_usd", 0.0) or 0.0)

        for extra_iterations in extra_answer_iterations:
            extra_iteration_count = max(0, int(extra_iterations))
            target_repeats = base_repeats + extra_iteration_count
            scaled_answer_cost = answer_cost * (target_repeats / base_repeats)
            estimated_total_cost = translation_cost + scaled_answer_cost
            communicated_budget = (
                math.ceil(estimated_total_cost / safe_round_increment)
                * safe_round_increment
            )

            if extra_iteration_count == 0:
                budget_tier = "Base plan"
            elif extra_iteration_count == 1:
                budget_tier = "+1 extra iteration"
            elif extra_iteration_count == recommended_extra_iterations:
                budget_tier = "Safe budget"
            else:
                budget_tier = f"Conservative (+{extra_iteration_count} extra iterations)"

            rows.append(
                {
                    "scenario_name": str(row.get("scenario_name", "")),
                    "budget_tier": budget_tier,
                    "extra_answer_iterations": extra_iteration_count,
                    "target_repeats_per_prompt": target_repeats,
                    "estimated_total_cost_usd": estimated_total_cost,
                    "communicated_budget_usd": communicated_budget,
                    "recommended_to_communicate": (
                        extra_iteration_count == recommended_extra_iterations
                    ),
                }
            )

    frame = pd.DataFrame(rows)
    frame["estimated_total_cost_usd"] = frame["estimated_total_cost_usd"].round(4)
    frame["communicated_budget_usd"] = frame["communicated_budget_usd"].round(2)
    return frame
